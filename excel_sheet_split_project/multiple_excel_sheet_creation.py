from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd

excel_files = [
    r'E:\Excel_django\excel_sheet_split_project\file1.xlsx',
    r'E:\Excel_django\excel_sheet_split_project\file2.xlsx',
    r'E:\Excel_django\excel_sheet_split_project\file3.xlsx'
]
output_excel = Workbook()
for excel_file in excel_files:
    df = pd.read_excel(excel_file)
    product_names = df['Product_Name'].unique()
    source_workbook = load_workbook(excel_file)
    for product_name in product_names:
        output_sheet = output_excel.create_sheet(title=str(product_name))
        for sheet in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet]
            for row in source_sheet.iter_rows(values_only=True):
                output_sheet.append(row)
output_excel.remove(output_excel.active)
output_excel.save('output_excel.xlsx')
print("Excel file 'output.xlsx' created successfully.")
