
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
import json
import pandas as pd
from my_app.serializers import FileUploadSerializer
from io import BytesIO
from django.http import HttpResponse,JsonResponse
from .models import Category,SubCategory
from .serializers import CategorySerializer,SubCategorySerializer
from .utils import read_document_and_extract_data,write_dicts_to_excel
from rest_framework.exceptions import ValidationError,NotFound
from django.db import transaction


class FileCombinedApi(APIView):
    def get(self, request, format=None):

        categories = Category.objects.all()
        serialized_data = []
        for category in categories:
            category_data = {
                "id":category.id,
                "category": category.category_name,
            }
            subcategories = SubCategory.objects.filter(category=category)
            for i, subcategory in enumerate(subcategories, start=1):
                category_data[f"sub category {i}"] = subcategory.sub_category_name
            
            serialized_data.append(category_data)
        
        return Response(serialized_data)
   

    def post(self, request):
        uploaded_datas=request.data.keys()
        uploaded_data1 = [uploaded_data for uploaded_data in uploaded_datas if uploaded_data.startswith('category')]
        file_datas = request.data.keys()
        file_data1 = [file_data for file_data in file_datas if file_data.startswith('file')]
        if uploaded_data1:
            return self.Add_Category(request)
        elif file_data1:
            return self.FileHandle(request)


    def Add_Category(self,request):
            data = request.data
            category_name = data.get('category_name')
            sub_category_names = data.get('sub_category_name')
            if not sub_category_names:
                if Category.objects.filter(category_name=category_name).exists():
                    return Response({"message": "This Category already exists"}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    category = Category.objects.create(category_name=category_name)
            else:
                try:
                    with transaction.atomic():
                        category = Category.objects.create(category_name=category_name)
                        for sub_category_name in sub_category_names:
                            SubCategory.objects.create(category=category, sub_category_name=sub_category_name)
                    return Response({"message": "Category and subcategories created successfully", "id": category.id}, status=status.HTTP_201_CREATED)
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
          

    def FileHandle(self,request):
        file_datas = request.data.keys()
        file_data1 = [file_data for file_data in file_datas if file_data.startswith('file')]
        if file_data1:
            file_subcategories = []
            for key in request.data.keys():
                if 'subcategory' not in key:
                    files = request.FILES.getlist(key)
                    subcategory_key = f"{key}_subcategory"
                    subcategory = request.data.get(subcategory_key)
                    for file in files:
                        file_subcategories.append({'file': file, 'subcategory': subcategory})

            serializer = FileUploadSerializer(data={'file_subcategories': file_subcategories})
            if serializer.is_valid():
                output_excel = Workbook()
                any_sheets = False
                all_data_dicts = []
                unique_subcategories = set(subcategory['subcategory'] for subcategory in file_subcategories)
                for subcategory in unique_subcategories:
                    subcategory_data_dicts = []
                    for file_subcategory in file_subcategories:
                        if file_subcategory['subcategory'] == subcategory:
                            excel_file = file_subcategory['file']
                            data_dicts,header =read_document_and_extract_data(excel_file)
                            subcategory_data_dicts.extend(data_dicts)
                            # subcategory_data_dicts.extend(extract_data_from_documents(excel_file))
                            
                    all_data_dicts.extend(subcategory_data_dicts)
                    output_sheet = output_excel.create_sheet(title=subcategory[:31])
                    headers = write_dicts_to_excel(subcategory_data_dicts).columns.tolist()
                    output_sheet.append(headers)
                    if not subcategory_data_dicts:
                        return JsonResponse({"error": f"No data found for subcategory: {subcategory}"}, status=status.HTTP_400_BAD_REQUEST)
                    df_filtered = write_dicts_to_excel(subcategory_data_dicts)
                    for _, row in df_filtered.iterrows():
                        row_values = row.tolist()
                        output_sheet.append(row_values)
                    any_sheets = True

                if any_sheets:
                    default_sheet = output_excel.active
                    output_excel.remove(default_sheet)

                    buffer = BytesIO()
                    output_excel.save(buffer)
                    buffer.seek(0)

                    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="combined.xlsx"'
                    return response
                else:
                    return JsonResponse({"error": "No data found for any subcategory"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                print(serializer.errors)
                return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        
    def delete(self, request, id=None, sub_category_name=[], format=None):

        if id and sub_category_name:
            breakpoint()
            return self.delete_sub_category_name(id, sub_category_name)
        elif id:
            return self.delete_category(id)
        else:
            return Response({"message": "Invalid request."}, status=status.HTTP_400_BAD_REQUEST)
        

    def delete_category(self,  id):

        try:
            category = Category.objects.get(id=id)
            subcategories = SubCategory.objects.filter(category=category)
            category.delete()
            subcategories.delete()
            return Response({"message": f"Category  and its subcategories deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except Category.DoesNotExist:
            return Response({"message": f"Category  does not exist"}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

    def delete_sub_category_name(self,id,sub_category_name):
        try:
            category = Category.objects.get(category_name=id)
            subcategories = SubCategory.objects.filter(category=category, sub_category_name__in=sub_category_name)
            subcategories.delete()
            return Response({"message": f"Subcategories '{', '.join(sub_category_name)}' in category  deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
        except Category.DoesNotExist:
            return Response({"message": f"Category  does not exist"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

    def put(self, request, id=None,category_name=None, format=None):

        if category_name and id:
            return self.update_category_and_subCategory(request,id)
        
        elif id:
             return self.update_category(request, id)
        
        else:
            return Response({"message": "Invalid request."}, status=status.HTTP_400_BAD_REQUEST)


    def update_category(self, request, id):

        try:
            categories = Category.objects.filter(id=id)
            if not categories.exists():
                return Response({'message': 'Category not Found'}, status=status.HTTP_400_BAD_REQUEST)
        except Category.MultipleObjectsReturned:
            return Response({'message': 'Duplicate category names found. Please fix data consistency.'}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        new_category_name = data.get('category')
        for category in categories:
            if new_category_name and new_category_name != category.category_name:
                category.category_name = new_category_name
                category.save()
            else:
                return Response({'message': 'Category was not updated!!!'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': 'Category was updated successfully'}, status=status.HTTP_200_OK)
    

    def update_category_and_subCategory(self, request, id):
        
        try:
            categories = Category.objects.filter(id=id)
            if not categories.exists():
                return Response({'message': 'Category not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Category.MultipleObjectsReturned:
            return Response({'message': 'Duplicate category names found. Please fix data consistency.'}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        new_category_name = data.get('category')
        sub_len = sum(1 for key in data.keys() if key.startswith('sub category'))
        for category in categories:
            if new_category_name and new_category_name != category.category_name:
                category.category_name = new_category_name
                category.save()

        for i in range(1,sub_len+1):
            sub_category_key = f'sub category {i}'
            sub_category_value = data.get(sub_category_key)
            if sub_category_value:
                subcategory, _ = SubCategory.objects.get_or_create(category=category, sub_category_name=sub_category_value)
                subcategory.sub_category_name = sub_category_value
                subcategory.save()
            else:
                return Response({'message': 'subcategories are not updated!!'},status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': 'Category and subcategories updated successfully'}, status=status.HTTP_200_OK)
    



















        # sub_len = sum(1 for key in data.keys() if key.startswith('sub category'))
        # sub_category_names = [value for key, value in data.items() if key.startswith('sub category')]
        # for category in categories:
        #     # if new_category_name and new_category_name != category.category_name:
        #     category.category_name = new_category_name
        #     subcategories = category.subcategory_set.all()
        #     for  i,subcategory in enumerate(subcategories[:sub_len]):
        #         print
        #         subcategory.sub_category_name = sub_category_names[i]
        #         category.save()
        #         subcategory.save()
        # else:
        #     return Response({'message': 'Check the category is already exists'})
        

    # # def update_category(self, request, category_name):
    # #     try:
    # #         categories = CategorySubcategory.objects.filter(category_name=category_name)
    # #         if not categories.exists():
    # #             return Response({'message': 'Category not Found'}, status=status.HTTP_400_BAD_REQUEST)
    # #     except CategorySubcategory.MultipleObjectsReturned:
    # #         return Response({'message': 'Duplicate category names found. Please fix data consistency.'}, status=status.HTTP_400_BAD_REQUEST)

    # #     data = request.data
    # #     new_category_name = data.get('category')
    # #     sub_category_names = [value for key, value in data.items() if key.startswith('sub category')]
    # #     breakpoint()
    # #     for category in categories:
    # #         if new_category_name and new_category_name != category_name:
    # #             category.category_name = new_category_name
    # #             for key, value in data.items():
    # #                 breakpoint()
    # #                 if key.startswith('sub category'):
    # #                     # Extract the subcategory number from the key
    # #                     sub_category_number = int(key.split()[-1])
    # #                     # Update the corresponding subcategory field in the database
    # #                     setattr(category, f'sub_category_{sub_category_number}', value)

    # #         # if sub_category_names:
    # #         #     breakpoint()
    # #         #     for sub_category_name in sub_category_names:
    # #         #         category.sub_category_name = sub_category_name
    # #             # for sub_category_name in sub_category_names:
    # #             #     category.sub_category_name.create(sub_category_name=sub_category_name)

    # #             category.save()

    # #     return Response({'message': 'Category and subcategories updated successfully'}, status=status.HTTP_200_OK)

        


    # def delete(self, request, category_name=None, sub_category_name=None, format=None):
    #     if category_name and sub_category_name:
    #         return self.delete_sub_category_name(category_name, sub_category_name)
    #     elif category_name:
    #         return self.delete_category(category_name)
    #     else:
    #         return Response({"message": "Invalid request."}, status=status.HTTP_400_BAD_REQUEST)

    # def delete_category(self,  category_name):
    #     categories = CategorySubcategory.objects.filter(category_name=category_name)
    #     if categories.exists():
    #         categories.delete()
    #         return Response({"message": "Category deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    #     else:
    #         return Response({"message": "Category not found."}, status=status.HTTP_404_NOT_FOUND)
        



        
    # def delete_sub_category_name(self,category_name,  sub_category_name):
    #     try:
    #         subcategory = CategorySubcategory.objects.get(sub_category_name=sub_category_name)
    #         # category_name = subcategory.category_name
    #         subcategory.delete()
    #         categories_with_subcategory = CategorySubcategory.objects.filter(category_name=category_name)
    #         if not categories_with_subcategory.exists():
    #             CategorySubcategory.objects.filter(category_name=category_name).delete()
    #         return Response({"message": "Sub-category deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    #     except CategorySubcategory.DoesNotExist:
    #         return Response({"message": "Sub-category not found."}, status=status.HTTP_404_NOT_FOUND)


















# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from django.http import HttpResponse
# from openpyxl import load_workbook, Workbook
# import pandas as pd
# from my_app.serializers import FileUploadSerializer
# from io import BytesIO




            # data = request.data
            # category_name = data.get('category_name')
            # checking_category = Category.objects.filter(category_name=category_name)
            # if checking_category.exists():
            #     return Response({"message": "Category already exists"}, status=status.HTTP_400_BAD_REQUEST)
            
            # sub_category_names = data.get('sub_category_name', [])
            # if len(sub_category_names) != len(set(sub_category_names)):
            #     raise ValidationError("Duplicate subcategory names are not allowed.")
            
            # categories = []
            # for sub_category_name in sub_category_names:
            #     categories.append({'category_name': category_name, 'sub_category_name': sub_category_name})

            # serialized_categories = []
            # for category_data in categories:
            #     serializer = CategorySerializer(data=category_data)
            #     breakpoint()
            #     if serializer.is_valid():
            #         serializer.save()
            #         serialized_categories.append(serializer.data)
            #     else:
            #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            # return Response(serialized_categories, status=status.HTTP_201_CREATED)

 # def get(self, request, category_name=None, format=None):
    #     if category_name:
    #         categories = CategorySubcategory.objects.filter(category_name=category_name)
    #         if categories.exists():
    #             serializer = CategorySerializer(categories, many=True)
    #             original_data = serializer.data
    #             transformed_data = {}
    #             for item in original_data:
    #                 category_name = item["category_name"]
    #                 sub_category_name = item["sub_category_name"]
    #                 if category_name in transformed_data:
    #                     transformed_data[category_name].append(sub_category_name)
    #                 else:
    #                     transformed_data[category_name] = [sub_category_name]
    #             final_data = []
    #             for category_name, sub_categories in transformed_data.items():
    #                 category_data = {"category": category_name}
    #                 for i, sub_category_name in enumerate(sub_categories, start=1):
    #                     category_data[f"sub category {i}"] = sub_category_name
    #                 final_data.append(category_data)
                    
    #             return Response(final_data)

    #         else:
    #             return Response({"message": "Category not found."}, status=status.HTTP_404_NOT_FOUND)
    #     else:
    #         categories = CategorySubcategory.objects.all()
    #         serializer = CategorySerializer(categories, many=True)
    #         original_data = serializer.data
    #         transformed_data = {}
    #         for item in original_data:
    #             category_name = item["category_name"]
    #             sub_category_name = item["sub_category_name"]
    #             if category_name in transformed_data:
    #                 transformed_data[category_name].append(sub_category_name)
    #             else:
    #                 transformed_data[category_name] = [sub_category_name]
    #         final_data = []
    #         for category_name, sub_categories in transformed_data.items():
    #             category_data = {"category": category_name}
    #             for i, sub_category_name in enumerate(sub_categories, start=1):
    #                 category_data[f"sub category {i}"] = sub_category_name
    #             final_data.append(category_data)
    #         return Response(final_data)
# class FileCombinedApi(APIView):
#     def post(self, request):
#         file_subcategories = []
#         for key in request.data.keys():
#             if not 'subcategory' in key:
#                 files = request.FILES.getlist(key)
#                 subcategory_key = f"{key}_subcategory"
#                 subcategory = request.data.get(subcategory_key)
#                 for file in files:
#                     file_subcategories.append({'file': file, 'subcategory': subcategory})

#         # request_data = {
#         #     'file_subcategories': file_subcategories,
#         #     # Include other required fields as needed
#         # }
#         # file1 = request.FILES.get('file1')
#         # # file2 = request.FILES.get('file2')

#         # # Other request data
#         # subcategory1 = request.data.get('subcategory1')
#         # request_data = {
#         #         'files': [file1],  # List of file objects
#         #         'subcategory1': subcategory1,
#         #         # Include other required fields as needed
#         #     }
#         serializer = FileUploadSerializer(data={'file_subcategories': file_subcategories})
#         if serializer.is_valid():
#             output_excel = Workbook()
#             any_sheets = False  # Flag to check if there are any sheets
#             for key, value in request.data.items():
#                 if key.startswith('file') and isinstance(value, list):
#                     subcategory_key = f"{key}_subcategory"
#                     subcategory = request.data.get(subcategory_key)
#                     subcategory=subcategory[:31]
#                     for file in value:
#                         df = pd.read_excel(file)
#                         output_sheet = output_excel.create_sheet(title=subcategory)
#                         headers = df.columns.tolist()
#                         output_sheet.append(headers)
#                         for _, row in df.iterrows():
#                             row_values = row.tolist()
#                             output_sheet.append(row_values)
#                         any_sheets = True  # Mark that at least one sheet is added

#             # Remove the default sheet only if there are other sheets
#             if any_sheets:
#                 default_sheet = output_excel.active
#                 output_excel.remove(default_sheet)

#             # Prepare the HTTP response with the output Excel file
#             buffer = BytesIO()
#             output_excel.save(buffer)
#             buffer.seek(0)

#             response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="combined.xlsx"'

#             return response
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#         serializer = FileUploadSerializer(data=request_data)
#         serializer = FileUploadSerializer(data=request.data)
#         if serializer.is_valid():
#             breakpoint()
#             excel_files = request.FILES.getlist('file1')
#             sheet_names= request.data
#             output_excel = Workbook()
#             for excel_file in excel_files:
#                 # breakpoint()
#                 df = pd.read_excel(excel_file)
#                 product_names = df['Product_Name'].unique()
#                 # product_names=df[df['Product_Name']==subcategory1]
#                 source_workbook = load_workbook(excel_file)
#                 for product_name in product_names:
#                     # product_name1=df[df['Product_Name']==subcategory1]
#                     breakpoint()
#                     # if product_name == subcategory1:
#                     output_sheet = output_excel.create_sheet(title=str(product_name))
#                     # df_filtered = df[df['Product_Name'] == subcategory1]
#                     headers = df.columns.tolist()
#                     output_sheet.append(headers)
#                     for index, row in df_filtered.iterrows():
#                         row_values = row.tolist()
#                         output_sheet.append(row_values)
#                         for sheet in source_workbook.sheetnames:
#                             breakpoint()
#                             source_sheet = source_workbook[sheet]
#                             for row in source_sheet.iter_rows(values_only=True):
#                                 breakpoint()
#                                 output_sheet.append(row)
#             Remove the default sheet created by output_excel

#             default_sheet = output_excel.active
#             output_excel.remove(default_sheet)

#             # Save the workbook to BytesIO buffer
#             buffer = BytesIO()
#             output_excel.save(buffer)

#             # Prepare the HTTP response with the output Excel file
#             response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="combined.xlsx"'

#             return response
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



    # def get(self, request, format=None):
    #     categories = Category.objects.all()
    #     serializer = CategorySerializer(categories, many=True).data
    #     temp_data = [
    #         {
    #             "category": "Things to See",
    #             "sub category 1": "Beaches",
    #             "sub category 2": "Parks, Zoos & Aquariums",
    #             "sub category 3": "Day Trips",
    #             "sub category 4": "Modern Tourist Attractions",
    #             "sub category 5": "Guided Sightseeing Tours",
    #             "sub category 6": "Museums",
    #             "sub category 7": "Art Galleries",
    #             "sub category 8": "Places of Worship",
    #         },
    #         {
    #             "category": "Things to Do",
    #             "sub category 1": "Amusement & Theme Parks",
    #             "sub category 2": "Kids Amusement Parks",
    #             "sub category 3": "Aqua Parks",
    #             "sub category 4": "Game Rooms",
    #             "sub category 5": "Boating & Cruises",
    #             "sub category 6": "Water-sports",
    #             "sub category 7": "Family-Friendly Adventure",
    #             "sub category 8": "Luxury Experiences",
    #             "sub category 8": "Wellness & Spa",
    #             "sub category 8": "Unique destination themes",
    #             "sub category 8": "Walking & Biking Tours",
    #             "sub category 8": "Guided Tours",
    #             "sub category 8": "Trekking"
    #         }
    #     ]
    #     data = serializer + temp_data
    #     return Response(data)




# data = request.data
        # # category_name=data.get('category_name')
        # # sub_category_name =data.get('sub_category_name',[])
        # # categories_pairs = []
        # # for i,sub in enumerate(sub_category_name,start=1):
        # #     categories_pairs.append({'category_name':category_name, 'sub_category_name':sub_category_name})
        # categories=[]

        # for key, value in data.items():
        #     category_name = value
        #     subcategory_key = key.replace('category', 'sub_category')
        #     subcategory_names = data.get(subcategory_key, [])
        #     for subcategory_name in subcategory_names:
        #         categories.append({'category_name':category_name,'sub_category_name':subcategory_name})
            
        # serialized_categories = []
        # for category_data in categories:
        #     breakpoint()
        #     serializer = CategorySerializer(data=category_data)
        #     if serializer.is_valid():
        #         serializer.save()
        #         serialized_categories.append(serializer.data)
        #     else:
        #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        #     return Response(serialized_categories, status=status.HTTP_201_CREATED)
        # uploaded_datas=request.data.keys()
        
        # uploaded_data1 = [uploaded_data for uploaded_data in uploaded_datas if uploaded_data.startswith('category')]
        # if uploaded_data1:

        #     serializer=CategorySerializer(data=request.data)
            
        #     if serializer.is_valid():
        #         serializer.save()
        #         return Response(serializer.data, status=status.HTTP_201_CREATED)
        #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            # breakpoint()
            # # required_keys=[request.data.get('category')]
            # sub_category_count=sum(1 for key in request.data.keys() if key.startswith('sub category') )
            # # required_keys.extend([f'sub category {i}' for i in range(sub_category_count)])
            # category = data["category"]
            # subcategories = [data[f"sub category {i}"] for i in range(1,(sub_category_count+1))]

